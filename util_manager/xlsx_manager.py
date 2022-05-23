import os
import openpyxl

from openpyxl.utils.cell import get_column_letter
from util_manager.list_manager import ListManager


class XlsxManager(object):
    """ 操作 Excel 的工具类 """

    def __init__(self):
        self.wb, self.sheet = None, None
        self.filepath, self.sheetname = '', ''
        self.list_manager = ListManager()

    def _create_excel(self, filepath):
        """ 返回全新的工作薄对象 """
        if not filepath.endswith('.xlsx'):
            raise ValueError('文件后缀不符')
        dirpath, _ = os.path.split(filepath)
        os.path.exists(dirpath) or os.makedirs(dirpath)
        self.wb = openpyxl.Workbook()
        self.wb.save(filepath)

    def _load_excel(self, filepath):
        """ 加载一个已经存在的工作薄 """
        if not filepath.endswith('.xlsx'):
            raise ValueError('文件后缀不符')
        self.wb = openpyxl.load_workbook(filepath)

    def init(self, filepath, sheetnames, create=False):
        """ 初始化工作薄，工作表 """
        self.filepath = filepath
        self._create_excel(filepath) if create else self._load_excel(filepath)
        self.new_and_clear_sheets(sheetnames)
        sheetname = self.wb.sheetnames[0]
        self.sheetname = sheetname
        self.sheet = self.wb[sheetname]

    def switch(self, filepath, sheetname, create=False):
        """ 切换工作薄和工作表 """
        if filepath != self.filepath:
            self._create_excel(filepath) if create else self._load_excel(filepath)
        if filepath != self.filepath or sheetname != self.sheetname:
            self.sheet = self._get_sheet(sheetname=sheetname)
        self.filepath, self.sheetname = filepath, sheetname

    def new_sheet(self, sheetname, index=-1):
        """ 在指定下标新建一个 sheet """
        if sheetname in self.wb.sheetnames: return
        if index < 0: index = len(self.wb.sheetnames)
        self.wb.create_sheet(sheetname, index)
        self._save()

    def new_sheets(self, sheetnames):
        """ 依次往后新建一群 sheet """
        for sheetname in sheetnames:
            self.new_sheet(sheetname)

    def new_and_clear_sheets(self, sheetnames):
        """ 新建指定的所有 sheet 同时删除旧的 sheet """
        self.new_sheets(sheetnames)
        diff = self.list_manager.together(self.wb.sheetnames, filter_=lambda x: x not in sheetnames)
        self.remove_sheets(diff)

    def remove_sheet(self, sheetname):
        """ 删除指定的一个 sheet """
        if sheetname not in self.wb.sheetnames: return
        del self.wb[sheetname]
        self._save()

    def remove_sheets(self, sheetnames):
        """ 删除指定的所有 sheet """
        for sheetname in sheetnames:
            self.remove_sheet(sheetname)

    def _get_sheet(self, sheetname):
        """ 从工作薄对象里获取工作表，不存在就创建 """
        if sheetname not in self.wb.sheetnames:
            self.wb.create_sheet(sheetname)
        return self.wb[sheetname]

    def _save(self):
        """ 保存修改 """
        self.wb.save(self.filepath)

    def size(self):
        """ 获取当前工作表内容最大行列数 """
        row, col = self.sheet.max_row, self.sheet.max_column
        if row == 1 and col == 1 and not self.sheet['A1'].value:
            row, col = 0, 0
        return row, col

    def read(self, start=None, end=None):
        """ 读取从表格数据，可以指定读取的范围 """
        results = []
        if not start:
            start = 'A1'
        if not end:
            end = f'{get_column_letter(self.sheet.max_column)}{self.sheet.max_row}'
        area = self.sheet[start: end]
        for line in area:
            tmp = []
            for cell in line:
                tmp.append(cell.value)
            results.append(tmp)
        return results

    def write(self, position, value):
        """ 往表格某个单元格里写数据 """
        self.sheet[position].value = value
        self._save()

    def write_line(self, row_number, data, need_save=True):
        """ 往表格指定行写一行数据 """
        for i, value in enumerate(data):
            self.sheet.cell(row=row_number, column=i+1).value = value
        if need_save: self._save()

    def write_lines(self, begin_row_number, datas):
        """ 往表格写很多行数据 """
        for i, data in enumerate(datas):
            self.write_line(begin_row_number + i, data, need_save=False)
        self._save()

    def append_line(self, data):
        """ 往表格追加一行数据 """
        row, _ = self.size()
        self.write_line(row + 1, data)

    def append_lines(self, datas):
        """ 往表格追加许多行数据 """
        row, _ = self.size()
        self.write_lines(row + 1, datas)


def main():
    path = '/Users/luyan/Desktop/example.xlsx'
    manager = XlsxManager()
    manager.switch(path, 'Sheet')
    print(manager.size())
    manager.append_lines([['小舞', '女', 19], ['奥斯卡', '男', 20]])


if __name__ == '__main__':
    main()
